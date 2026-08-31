from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import User, Property, Agreement, Payment

from openpyxl import load_workbook

class ImportUsers(APIView):
    def post(self, request):
        file = request.FILES.get("file")

        spreadsheet = load_workbook(file)

        sheet = spreadsheet["Usuarios"]

        created = 0
        ignored = 0

        for column in sheet.iter_rows(min_row=2, values_only=True):
            username = column[0]
            first_name = column[1]
            last_name = column[2]
            email = column[3]
            telephone = column[4]
            user_type = column[5]
            is_active = column[6]
            password = column[7]

            if User.objects.filter(username=username).exists():
                ignored += 1
                continue

            user = User(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                telephone=telephone,
                user_type=user_type,
                is_active=is_active,
                password=password
            )

            user.set_password(password)
            user.save()

            created += 1

        return Response(
            {
                "message": "Planilha lida com sucesso!",
                "created_users": created,
                "ignored_users": ignored
            },
            status=status.HTTP_201_CREATED
        )


class ImportPropertys(APIView):
    def post(self, request):
        file = request.FILES.get("file")
        spreadsheet = load_workbook(file)
        sheet = spreadsheet["Imoveis"]

        created = 0
        ignored = 0

        for column in sheet.iter_rows(min_row=2, values_only=True):
            [ 
                title, property_type, 
                rent_price, property_status, 
                address, cep, 
                complement, neighborhood, city, uf 
            ] = column


            if Property.objects.filter(address=address).exists():
                ignored += 1
                continue

            newProperty = Property(
                title=title,
                property_type=property_type,
                rent_price=rent_price,
                status=property_status,
                address=address,
                cep=cep,
                complement=complement,
                neighborhood=neighborhood,
                city=city,
                uf=uf
            )

            newProperty.save()
            created += 1


        return Response(
            {
                "message": "Planilha lida com sucesso!",
                "created_propertys": created,
                "ignored_propertys": ignored
            },
            status=status.HTTP_201_CREATED
        )


class ImportAgreements(APIView):
    def post(self, request):
        file = request.FILES.get("file")
        spreadsheet = load_workbook(file)
        sheet = spreadsheet["Contratos"]

        created = 0

        for column in sheet.iter_rows(min_row=2, values_only=True):
            [  
                start_date,
                end_date,
                price,
                agreement_status,
                lessor_id,
                renter_id,
                ignored_field
            ] = column

            newAgreement = Agreement(
                start_date=start_date,
                end_date=end_date,
                price=price,
                status=agreement_status,
                lessor_id=lessor_id,
                renter_id=renter_id
            )

            newAgreement.save()
            created += 1
        

        return Response(
            {
                "message": "Planilha lida com sucesso!",
                "created_agreements": created
            },
            status=status.HTTP_201_CREATED
        )


class ImportPayments(APIView):
    def post(self, request):
        file = request.FILES.get("file")
        spreadsheet = load_workbook(file)
        sheet = spreadsheet["Pagamentos"]

        created = 0
        ignored = 0


        for column in sheet.iter_rows(min_row=2, values_only=True):
            [
                payment_date,
                price,
                payment_status,
                agreement_id,
                ignored_field
            ] = column

            if Payment.objects.filter(agreement_id=agreement_id).exists():
                ignored += 1
                continue

            newPayment = Payment(
                payment_date=payment_date,
                price=price,
                status=payment_status,
                agreement_id=agreement_id
            )

            newPayment.save()
            created += 1


        return Response(
            {
                "message": "Planilha lida com sucesso!",
                "created_payments": created,
                "ignored_payments": ignored
            },
            status=status.HTTP_201_CREATED
        )